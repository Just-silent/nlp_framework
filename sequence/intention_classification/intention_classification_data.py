# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/9/18 8:16
import pandas as pd
import torch
import random
import numpy as np
from transformers import BertTokenizer

import openpyxl


class IntentionClassificationDataLoader(object):
    def __init__(self, config):
        # , data_dir, bert_class, params, token_pad_idx=0, tag_pad_idx=-1
        self._config = config
        self.batch_size = config.data.batch_size
        self.max_len = config.data.max_len
        self.token_pad_idx = config.data.token_pad_idx
        self.tag_pad_idx = config.data.tag_pad_idx
        self._shuffle = config.data.shuffle
        self.device = config.device
        self.seed = config.project.seed

        tags = self.load_tags()
        self.tag2idx = {tag: idx for idx, tag in enumerate(tags)}
        self.idx2tag = {idx: tag for idx, tag in enumerate(tags)}

        self.tokenizer = BertTokenizer.from_pretrained(config.pretrained_models.dir, do_lower_case=False)

        self.train_data = {}
        self.valid_data = {}
        self.test_data = {}

        self.load_data(config.data.train_path, self.train_data)
        self.load_data(config.data.valid_path, self.valid_data)
        self.load_data(config.data.test_path, self.test_data)

    def load_tags(self):
        tags = []
        data = pd.read_excel(self._config.data.train_path)
        ori_tags = list(data['二级分类'].values)
        for tag in ori_tags:
            if tag is not np.nan:
                tag = tag.strip()
                tags.append(tag)
        result = []
        if self._config.model.label_pad:
            result = ['PAD']
        tags_new = list(set(tags))
        tags_new.sort(key=tags.index)
        result.extend(tags_new)
        return result
        # return list(set(tags))

    #将原始数据切分为train和valid
    def cut_data(self):
        all_data = {}
        tags = []
        wb = openpyxl.load_workbook(self._config.data.ori_path)
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]
        wb_train = openpyxl.load_workbook(self._config.data.train_path)
        ws_train = wb_train.create_sheet(sheetnames[0])
        ws_train = wb_train[sheetnames[0]]
        wb_valid = openpyxl.load_workbook(self._config.data.valid_path)
        ws_valid = wb_valid.create_sheet(sheetnames[0])
        ws_valid = wb_valid[sheetnames[0]]
        max_row = ws.max_row
        for i in range(max_row-2):
            line_num = i+2
            tags.append(ws.cell(line_num, 5).value.strip())
        tags = list(set(tags))
        for i in range(len(tags)):
            all_data[tags[i]] = []
        for i in range(max_row-2):
            line_num = i+2
            all_data[ws.cell(line_num, 5).value.strip()].append(ws.cell(line_num, 3).value)
        line_train = 2
        line_valid = 2
        for key, vs in zip(all_data.keys(), all_data.values()):
            if len(vs)>1:
                for v in vs[:-1]:
                    ws_train.cell(line_train, 5).value = key
                    ws_train.cell(line_train, 3).value = v
                    line_train+=1
                ws_valid.cell(line_valid, 5).value = key
                ws_valid.cell(line_valid, 3).value = vs[-1]
                line_valid += 1
            else:
                ws_train.cell(line_train, 5).value = key
                ws_train.cell(line_train, 3).value = vs[0]
                line_train += 1
        wb_train.save(self._config.data.train_path)
        wb_valid.save(self._config.data.valid_path)
        pass

    def load_sentences_tags(self, path, data):
        """
        Args:
            path : the path of the data_file
            data : the list of data
        Returns:
            the list of all vocabs
        """
        sentences = []
        tags = []
        wb = openpyxl.load_workbook(path)
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]
        max_row = ws.max_row
        for i in range(max_row - 2):
            line_num = i + 2
            if ws.cell(line_num, 2).value is not None and ws.cell(line_num, 3).value is not None:
                # if ws.cell(line_num, 5).value is not None:
                sentence = ws.cell(line_num, 3).value.strip()
                # sentence = ws.cell(line_num, 5).value.strip()
                # if ws.cell(line_num, 5).value is not None:
                #     sentence = ws.cell(line_num, 3).value.strip() + ws.cell(line_num, 5).value.strip()
                tag = self.tag2idx[ws.cell(line_num, 2).value.strip()]
                subwords = list(map(self.tokenizer.tokenize, sentence))
                subword_lengths = list(map(len, subwords))
                subwords = ['[CLS]'] + [item for indices in subwords for item in indices]
                token_start_idxs = 1 + np.cumsum([0] + subword_lengths[:-1])
                sentences.append((self.tokenizer.convert_tokens_to_ids(subwords), token_start_idxs))
                tags.append(tag)
        data['tags'] = tags
        data['data'] = sentences
        data['size'] = len(sentences)

    def load_data(self, path , data):
        """Loads the data for each type in types from data_dir.

        Args:
            data_type: (str) has one of 'train', 'val', 'test' depending on which data is required.
        Returns:
            data: (dict) contains the data with tags for each type in types.
        """
        self.load_sentences_tags(path, data)

    def data_iterator(self, data):
        """Returns a generator that yields batches data with tags.

        Args:
            data: (dict) contains data which has keys 'data', 'tags' and 'size'
            shuffle: (bool) whether the data should be shuffled

        Yields:
            batch_data: (tensor) shape: (batch_size, max_len)
            batch_tags: (tensor) shape: (batch_size, max_len)
        """

        # make a list that decides the order in which we go over the data- this avoids explicit shuffling of data
        order = list(range(data['size']))
        if self._shuffle:
            random.seed(self.seed)
            random.shuffle(order)

        interMode = False if 'tags' in data else True

        if data['size'] % self.batch_size == 0:
            BATCH_NUM = data['size'] // self.batch_size
        else:
            BATCH_NUM = data['size'] // self.batch_size + 1

        # one pass over data
        for i in range(BATCH_NUM):
            # fetch sentences and tags
            if i * self.batch_size < data['size'] < (i + 1) * self.batch_size:
                sentences = [data['data'][idx] for idx in order[i * self.batch_size:]]
                if not interMode:
                    tags = [data['tags'][idx] for idx in order[i * self.batch_size:]]
            else:
                sentences = [data['data'][idx] for idx in order[i * self.batch_size:(i + 1) * self.batch_size]]
                if not interMode:
                    tags = [data['tags'][idx] for idx in order[i * self.batch_size:(i + 1) * self.batch_size]]

            # batch length
            batch_len = len(sentences)

            # compute length of longest sentence in batch
            batch_max_subwords_len = max([len(s[0]) for s in sentences])
            max_subwords_len = min(batch_max_subwords_len, self.max_len)
            max_token_len = 0

            # prepare a numpy array with the data, initialising the data with pad_idx
            batch_data = self.token_pad_idx * np.ones((batch_len, max_subwords_len))
            batch_token_starts = []

            # copy the data to the numpy array
            for j in range(batch_len):
                cur_subwords_len = len(sentences[j][0])
                if cur_subwords_len <= max_subwords_len:
                    batch_data[j][:cur_subwords_len] = sentences[j][0]
                else:
                    batch_data[j] = sentences[j][0][:max_subwords_len]
                token_start_idx = sentences[j][-1]
                token_starts = np.zeros(max_subwords_len)
                token_starts[[idx for idx in token_start_idx if idx < max_subwords_len]] = 1
                batch_token_starts.append(token_starts)
                max_token_len = max(int(sum(token_starts)), max_token_len)

            if not interMode:
                batch_tags = self.tag_pad_idx * np.ones((batch_len))
                for j in range(batch_len):
                    batch_tags[j] = tags[j]

            # since all data are indices, we convert them to torch LongTensors
            batch_data = torch.tensor(batch_data, dtype=torch.long)
            batch_token_starts = torch.tensor(batch_token_starts, dtype=torch.long)
            if not interMode:
                batch_tags = torch.tensor(batch_tags, dtype=torch.long)

            # shift tensors to GPU if available
            batch_data, batch_token_starts = batch_data.to(self.device), batch_token_starts.to(self.device)
            if not interMode:
                batch_tags = batch_tags.to(self.device)
                yield batch_data, batch_token_starts, batch_tags
            else:
                yield batch_data, batch_token_starts

    def load_train(self):
        return self.train_data

    def load_valid(self):
        return self.valid_data

    def load_test(self):
        return self.test_data


if __name__ == '__main__':
    config_file = 'intention_classification_config.yml'
    import dynamic_yaml

    # Device
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")

    with open(config_file, mode='r', encoding='UTF-8') as f:
        config = dynamic_yaml.load(f)
    data_loader = IntentionClassificationDataLoader(config)
    # datas = data_loader.cut_data()
    pass