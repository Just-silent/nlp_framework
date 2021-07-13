# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/9/18 8:49

from openpyxl import load_workbook
from torchtext.data import Dataset, Example

from sequence.emo.utils import Tool


class EmoDataset(Dataset):

    def __init__(self, path, fields, file, config, **kwargs):
        self._config = config
        self._tool = Tool()
        examples = self._get_examples(path, fields, file)
        super(EmoDataset, self).__init__(examples, fields, **kwargs)


    def _get_examples(self, path, fields, file):
        examples = []
        wb = load_workbook(path)
        wb_names = wb.sheetnames
        ws = wb[wb_names[0]]
        index_row = 2
        text_col = 5
        tag_col = 3
        max_row = ws.max_row
        while index_row<=max_row:
            text = ws.cell(index_row, text_col).value
            text_list = text.split()
            tag_list = int(ws.cell(index_row, tag_col).value)
            index_row+=1
            examples.append(Example.fromlist([text_list, tag_list], fields))
        return examples