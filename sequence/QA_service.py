# coding:UTF-8
# author    :Just_silent
# init time :2021/4/26 19:59
# file      :QA_service.py
# IDE       :PyCharm

import openpyxl
from flask import *
from time import time
from tqdm import tqdm
from py2neo import Graph, Node, Relationship, NodeMatcher, RelationshipMatcher, Subgraph
from sequence.bert_ner.bert_ce_runner import Bert_Runner
from sequence.intention_classification.intention_classification_runner import IntentionClassificationRunner


class Neo4jImport():
    def __init__(self):
        self.graf = Graph(
            "http://172.22.179.237:7474/",
            user="neo4j",
            password="123456"
        )
        self.email_dict = {}
        self.macth_node('普通附件', '添加')
        self.node_match = NodeMatcher(self.graf)
        self.rel_match = RelationshipMatcher(self.graf)
        self.icr = IntentionClassificationRunner('intention_classification_config.yml')
        self.icr.train()
        self.ner = Bert_Runner('bert_ce_config.yml')
        self.ner.train()
        pass

    def create_node(self, names, label):
        nodes = []
        for name in tqdm(names):
            if list(self.node_match.match(label, name=name))==[]:
                nodes.append(Node(label, name=name))
        if nodes!=[]:
            subgraf = Subgraph(nodes)
            self.graf.create(subgraf)
            print('完成创建节点{}个'.format(len(nodes)))
        pass

    def create_rel(self, rel_ds):
        relation_ships = []
        for rel_d in rel_ds:
            node1 = None
            node2 = None
            rel = rel_d['rel']
            if  list(self.node_match.match(rel_d['label1'], name=rel_d['name1']))!=[]:
                node1=list(self.node_match.match(rel_d['label1'], name=rel_d['name1']))[0]
            else:
                node1=Node(rel_d['label1'], name=rel_d['name1'])
            if  list(self.node_match.match(rel_d['label2'], name=rel_d['name2']))!=[]:
                node2=list(self.node_match.match(rel_d['label2'], name=rel_d['name2']))[0]
            else:
                node2=Node(rel_d['label2'], name=rel_d['name2'])
            relation_ship = Relationship(node1, rel, node2)
            relation_ships.append(relation_ship)
        if relation_ships!=[]:
            subgraf = Subgraph(relationships=relation_ships)
            self.graf.create(subgraf)
            print('完成创建关系{}个'.format(len(relation_ships)))
        pass

    def macth_node(self, keyword, intent):
        sql = 'match (:keyword{name:\'' + keyword + '\'})-[:`' + intent + '`]->(m) return m'
        result = self.graf.run(sql).data()
        return result

    def create_email(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        max_row = ws.max_row
        for line in range(2, max_row+1):
            domain1 = ws.cell(line, 1).value
            domain2 = ws.cell(line, 2).value
            question = ws.cell(line, 3).value
            keyword = ws.cell(line, 4).value
            relation = ws.cell(line, 5).value
            solution = ws.cell(line, 6).value
            if domain1 not in self.email_dict.keys():
                self.email_dict[domain1] = {}
                if domain2 not in self.email_dict[domain1].keys():
                    self.email_dict[domain1][domain2] = []
                    d = {}
                    d['question'] = question
                    d['keyword'] = keyword
                    d['relation'] = relation
                    d['solution'] = solution
                    self.email_dict[domain1][domain2].append(d)
                else:
                    d = {}
                    d['question'] = question
                    d['keyword'] = keyword
                    d['relation'] = relation
                    d['solution'] = solution
                    self.email_dict[domain1][domain2].append(d)
            elif domain2 not in self.email_dict[domain1].keys():
                self.email_dict[domain1][domain2] = []
                d = {}
                d['question'] = question
                d['keyword'] = keyword
                d['relation'] = relation
                d['solution'] = solution
                self.email_dict[domain1][domain2].append(d)
            else:
                d = {}
                d['question'] = question
                d['keyword'] = keyword
                d['relation'] = relation
                d['solution'] = solution
                self.email_dict[domain1][domain2].append(d)
        domain0 = ['内网邮件']
        d0rd1 = []
        domain1s = []
        d1rd2 = []
        domain2s = []
        d2rk = []
        keywords = []
        krqs = []
        q_s = []
        domain1s = self.email_dict.keys()
        for domain in domain1s:
            d = {'name1':domain0[0],
                 'label1':'domain0',
                 'rel':'domain02domain1',
                 'name2': domain,
                 'label2': 'domain1'
                 }
            d0rd1.append(d)
        for key in self.email_dict:
            for domain2 in self.email_dict[key].keys():
                domain2s.append(domain2)
                d = {
                    'name1': key,
                    'label1': 'domain1',
                    'rel': 'domain12domain2',
                    'name2': domain2,
                    'label2': 'domain2'
                }
                d1rd2.append(d)
                for d1 in self.email_dict[key][domain2]:
                    d = {
                        'name1': domain2,
                        'label1': 'domain2',
                        'rel': 'domain22keyword',
                        'name2': d1['keyword'],
                        'label2': 'keyword'
                    }
                    d2rk.append(d)
                    keywords.append(d1['keyword'])
                    q_s.append(d1['solution'])
                    d = {
                        'name1': d1['keyword'],
                        'label1': 'keyword',
                        'rel': d1['relation'],
                        'name2': d1['solution'],
                        'label2': 'solution'
                    }
                    krqs.append(d)
        domain1s = list(set(domain1s))
        domain2s = list(set(domain2s))
        keywords = list(set(keywords))
        q_s = list(set(q_s))

        self.create_node(domain0, label='domain0')
        self.create_node(domain1s, label='domain1')
        self.create_node(domain2s, label='domain2')
        self.create_node(keywords, label='keywords')
        self.create_node(q_s, label='solution')

        self.create_rel(d0rd1)
        self.create_rel(d1rd2)
        self.create_rel(d2rk)
        self.create_rel(krqs)


        pass


app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False

import sys
sys.path.append(r'E:\Github\nlp_framework\sequence\bert_ner')

neo4j = Neo4jImport()

@app.route('/')
def index():
    return render_template('main.html')


@app.route('/info', methods=['GET'])
def email():
    # 获取文本
    text = request.args.get("cw_question")
    print(text)
    keywords = neo4j.ner.predict_test(text)[0]
    intent = neo4j.icr.predict_test(text)
    result = neo4j.macth_node(keywords, intent)
    return render_template('main.html', questuon=text, result=result)



@app.errorhandler(500)
def error(e):
    return '错误，请检查代码！'


if __name__ == '__main__':
    print(app.url_map)
    # app.run(port=7777, debug=False, host='0.0.0.0')
    app.run(port=7777, debug=False, host='127.0.0.1')
    # app.run(port=7777, debug=False)



# data_path = r'C:\Users\Administrator\Desktop\运维知识整理-内网邮件\邮箱用户手册整理.xlsx'
# neo4j_import = Neo4jImport()
# neo4j_import.create_email(data_path)