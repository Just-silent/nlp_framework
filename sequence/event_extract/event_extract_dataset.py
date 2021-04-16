# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/9/18 8:49

from openpyxl import load_workbook
from torchtext.data import Dataset, Example

from sequence.event_extract.utils import Tool


class EEDataset(Dataset):
    def __init__(self, path, fields, file, config, **kwargs):
        self._config = config
        self._tool = Tool()
        examples = self._get_examples(path, fields, file)
        super(EEDataset, self).__init__(examples, fields, **kwargs)

    def _get_examples(self, path, fields, file):
        examples = []
        wb = load_workbook(path)
        wb_names = wb.sheetnames
        ws = wb[wb_names[0]]
        index_row = 2
        text_col = 5
        tag_col = 11
        max_row = ws.max_row
        while index_row<=max_row:
            text = ws.cell(index_row, text_col).value
            text_list = [c for c in text]
            words = ws.cell(index_row, tag_col).value.split()
            tag_list = self._tool.get_tags(text, words)
            index_row+=1
            examples.append(Example.fromlist([text_list, tag_list], fields))
        return examples